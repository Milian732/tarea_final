#!/usr/bin/python3

import os
import sys
import openpyxl

print(" * Welcome to the Diff Excel User System *")

vCampos=["login","Nombre","Telefono","Correo"]

ruta_ayer = sys.argv[1]
ruta_hoy = sys.argv[2]

print(" - Viejo fue : "+str(ruta_ayer))
print(" - Nuevo fue : "+str(ruta_hoy))

wb_ayer = openpyxl.load_workbook(ruta_ayer)
ws_ayer = wb_ayer.active

wb_hoy = openpyxl.load_workbook(ruta_hoy)
ws_hoy = wb_hoy.active

if os.path.exists("meta-script.sh"):
    print(" * Meta Script Cleaning System ")
    os.remove("meta-script.sh")

if os.path.exists("Contratos.md"):
    print(" * Contratos MarkDown Cleaning System ")
    os.remove("Contratos.md")

meta_script = open("meta-script.sh",'x')
contratos = open("Contratos.md",'x')
meta_script.write("#!/bin/bash\n\n")
meta_script.write("# DO NOT EDIT THIS SCRIPT \n")
meta_script.write("# IT WILL BE AUTOMAGICALLY GENERATED\n\n")
meta_script.write("# Orig "+ruta_ayer+"\n")
meta_script.write("# Dest "+ruta_hoy+"\n")
contratos.write("## TAREA FINAL JENKINS\n")
contratos.write("\n### FICHEROS DE LOS USUARIOS \n\n")
contratos.write("Viejo: "+ruta_ayer+"\n")
contratos.write("\nNuevo: "+ruta_hoy+"\n")

def nuevosContratos():
    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1
    
    meta_script.write("\n# NUEVOS CONTRATOS : \n")
    contratos.write("\n\n### NUEVOS CONTRATOS : \n\n")
    while (aux_id_hoy != None):
        
        aux_id_ayer=ws_ayer.cell(row=1,column=1).value
        fila_ayer_procesada = 1
        contrato = True
        
        while (aux_id_ayer != None):
            if aux_id_ayer == ws_hoy.cell(row=fila_hoy_procesada,column=1).value : 
                contrato = False
            # Siguiente linea mecanismo 
            fila_ayer_procesada = fila_ayer_procesada + 1
            aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value
        
        if contrato :
            auxUser = ws_hoy.cell(row=fila_hoy_procesada,column=2).value
            auxFull = ws_hoy.cell(row=fila_hoy_procesada,column=3).value
            auxTel = ws_hoy.cell(row=fila_hoy_procesada,column=4).value
            auxMail = ws_hoy.cell(row=fila_hoy_procesada,column=5).value
            auxUID = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
            print(" * Contrato a "+ws_hoy.cell(row=fila_hoy_procesada,column=3).value )
            
            meta_script.write("useradd -m -d \"/home/"+auxUser+"\" -s \"/bin/bash\" -u "+str(auxUID)+" -c \""+auxFull+", ,"+str(auxTel)+", ,"+auxMail+"\" \""+auxUser+"\"\n" )
            meta_script.write("echo \""+auxUser+":"+str(auxTel)+"\"| chpasswd \n")
            contratos.write("Se ha contratado a "+auxUser+"\n")
        
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value

def modificaciones():

    aux_id_hoy=ws_hoy.cell(row=1,column=1).value
    fila_hoy_procesada = 1

    meta_script.write("\n\n# MODIFICACIONES : \n")
    contratos.write("\n\n### MODIFICACIONES : \n")

    while (aux_id_hoy != None):
        
        old_fila = 1
        old_id = ws_ayer.cell(row=old_fila,column=1).value
       
        vCambiosUsuario = []
        
        while (old_id != None):
            if old_id == aux_id_hoy:
                # fila_ayer = old_fila
                # fila_hoy es fila_hoy_procesada
                hayCambios = False
                for campo in range(2,6):
                    if ws_ayer.cell(row=old_fila,column=campo).value != ws_hoy.cell(row=fila_hoy_procesada,column=campo).value:
                        vCambiosUsuario.append(campo-2)
                        hayCambios = True
                
                if hayCambios :
                    print(ws_hoy.cell(row=fila_hoy_procesada,column=2).value + " ha cambiado : ")
                    print(vCambiosUsuario)
                    for campoCambiado in range(0,len(vCambiosUsuario)):
                        auxC = int(vCambiosUsuario[campoCambiado])+2
                        print("  --->   "+vCampos[vCambiosUsuario[campoCambiado]]+" : "+ ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)
                    #    print(" * "+vCampos[campoCambiado])
                        print()

                        if vCampos[vCambiosUsuario[campoCambiado]] == "Nombre" :
                            meta_script.write("\nchfn -f \""+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+"\" "+ws_ayer.cell(row=old_fila,column=2).value+"\n")
                            contratos.write("\nSe ha modificado el campo "+vCampos[vCambiosUsuario[campoCambiado]]+" de "+ws_ayer.cell(row=old_fila,column=2).value+" a "+ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+"\n")
                        if vCampos[vCambiosUsuario[campoCambiado]] == "Telefono" :
                            meta_script.write("\nchfn -w \""+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+"\" "+ws_ayer.cell(row=old_fila,column=2).value+"\n")
                            contratos.write("\nSe ha modificado el campo "+vCampos[vCambiosUsuario[campoCambiado]]+" de "+ws_ayer.cell(row=old_fila,column=2).value+" a "+ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+"\n")
                        if vCampos[vCambiosUsuario[campoCambiado]] == "Correo" : 
                            meta_script.write("\nchfn -o \""+ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+"\" "+ws_ayer.cell(row=old_fila,column=2).value+"\n")
                            contratos.write("\nSe ha modificado el campo "+vCampos[vCambiosUsuario[campoCambiado]]+" de "+ws_ayer.cell(row=old_fila,column=2).value+" a "+ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+"\n")
                        if vCampos[vCambiosUsuario[campoCambiado]] == "login" :
                            meta_script.write("\nusermod -l "+str(ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value)+" "+ws_ayer.cell(row=old_fila,column=2).value+"\n")
                            contratos.write("\nSe ha modificado el campo "+vCampos[vCambiosUsuario[campoCambiado]]+" de "+ws_ayer.cell(row=old_fila,column=2).value+" a "+ws_hoy.cell(row=fila_hoy_procesada,column=auxC).value+"\n")


            old_fila = old_fila + 1
            old_id = ws_ayer.cell(row=old_fila,column=1).value
        
        # Siguiente linea mecanismo 
        fila_hoy_procesada = fila_hoy_procesada + 1
        aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value


def despidos():

    meta_script.write("\n\n# DESPIDOS PROCEDENTES :  \n\n")
    contratos.write("\n\n### DESPIDOS PROCEDENTES : \n")

    aux_id_ayer=ws_ayer.cell(row=1,column=1).value
    fila_ayer_procesada = 1

    while (aux_id_ayer != None):
        
        aux_id_hoy=ws_hoy.cell(row=1,column=1).value
        fila_hoy_procesada = 1
        despido = True
        
        while (aux_id_hoy != None):
            if aux_id_hoy == ws_ayer.cell(row=fila_ayer_procesada,column=1).value : 
                despido = False
            # Siguiente linea mecanismo 
            fila_hoy_procesada = fila_hoy_procesada + 1
            aux_id_hoy = ws_hoy.cell(row=fila_hoy_procesada,column=1).value
        
        if despido :
            print(" * Despide a "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value )
            
            meta_script.write("# Deleting "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value+"\n")
            meta_script.write("deluser "+ws_ayer.cell(row=fila_ayer_procesada,column=2).value+"\n")
            meta_script.write("\n")
            
            contratos.write("\nSe ha despedido a "+ws_ayer.cell(row=fila_ayer_procesada,column=3).value+"\n")
        
        #for columna in range(2,6):
        #    print(ws_ayer.cell(row=fila_ayer_procesada,column=columna).value)
            
        # Siguiente linea mecanismo 
        fila_ayer_procesada = fila_ayer_procesada + 1
        aux_id_ayer = ws_ayer.cell(row=fila_ayer_procesada,column=1).value

despidos()
nuevosContratos()        
modificaciones()    

# Print exit 0
meta_script.write("exit 0\n")
contratos.close()
meta_script.close()

sys.exit(0)





